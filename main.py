# Work with Python 3.6.5
import discord
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from random import *
import random
from PIL import Image, ImageDraw, ImageFont
import textwrap

def getConfig(fileName, command):
    """GET THE TOKEN FROM THE CONFIG FILE"""
    configFile = open(fileName, "r")
    linelist = configFile.read().splitlines()
    for line in linelist:
        if not line.startswith("#"):
            split = line.split("=")
            # TOKEN
            if split[0] == command:
                return split[1]
    return command + " <<< Command Not Found."

def readPermissions():
    # Check if the log file already exists, otherwise create one to write and read from.
    if not os.path.isfile('./permission.txt'):
        permissionFile = open("permission.txt", 'w')
        permissionFile.close()

    permissionFile = open("permission.txt", 'r')
    permissions = permissionFile.read()
    permissionFile.close()
    return permissions

def setPermissions(newRole):
    # Read The file
    logFile = open("permission.txt", "r")
    roles = logFile.read().split(";")
    logFile.close()

    # Turn it into a list
    roleList = []
    for role in roles:
        roleList.append(role)

    # Add all of the Roles to a string
    roleString = ""
    for role in roleList:
        roleString = roleString + role + ";"
    roleString = roleString + str(newRole)

    # Add to Permission List
    logFile = open("permission.txt", "w")
    logFile.write(roleString)
    logFile.close()

    return "Permission added for role: " + newRole + ". The permissions are now: " + readPermissions()

def readLogChannel():
    # Check if the log file already exists, otherwise create one to write and read from.
    if not os.path.isfile('./logchannel.txt'):
        logFile = open("logchannel.txt", 'w')
        logFile.close()
        return "Log channel made."

    logFile = open("logchannel.txt", 'r')
    log_channel = logFile.read()
    LOG_CHANNEL_ID = log_channel[2:-1]

    list = []
    list.append(log_channel)
    list.append(LOG_CHANNEL_ID)
    logFile.close()
    return list

def readDefaultChannel():
    # Check if the log file already exists, otherwise create one to write and read from.
    if not os.path.isfile('./def_channel.txt'):
        logFile = open("def_channel.txt", 'w')
        logFile.close()
        return "Default channel made."

    logFile = open("def_channel.txt", 'r')

    DEFAULT_CHANNEL = logFile.read()
    DEFAULT_CHANNEL_ID = DEFAULT_CHANNEL[2:-1]

    list = []
    list.append(DEFAULT_CHANNEL)
    list.append(DEFAULT_CHANNEL_ID)

    logFile.close()
    return list

def setLogChannel(log_channel):
    logFile = open("logchannel.txt", "w")
    logFile.write(log_channel)
    logFile.close()
    return "Log Channel is now: " + readLogChannel()[0]

def setDefaultChannel(def_channel):
    logFile = open("def_channel.txt", "w")
    logFile.write(def_channel)
    logFile.close()
    return "def Channel is now: " + def_channel

def getPermittedUserList():
    # Seek who in the discord is able to use the bot with his or her current role.
    memberList = []
    for member in client.get_all_members():
        memberList.append(member)

    membersWithPermission = []
    for member in memberList:
        permissionRoles = readPermissions()
        for roles in member.roles:
            if str(roles) in permissionRoles:
                membersWithPermission.append(member.id)
                break
    return membersWithPermission

def isPermitted(user):
    permissionList = getPermittedUserList()
    if user in permissionList:
        return True
    return False

def addGame(gamesAndKey):
    try:
        book = load_workbook('keys.xlsx')
        sheet = book.active
        lastKnownCell = None
        for row in sheet:
            for col in row:
                if col.value != None:
                    lastKnownCell = col

        newGameCell = str(chr(int(lastKnownCell.col_idx) + 63)) + str(int(lastKnownCell.row) + 1)
        newKeyCell = str(chr(int(lastKnownCell.col_idx) + 64)) + str(int(lastKnownCell.row) + 1)

        if gamesAndKey[0].startswith("<"):
            for string in gamesAndKey[1:]:
                if string.endswith(">"):
                    gameName = gamesAndKey[0][1:-1]
                    break

        for string in gamesAndKey[1:]:
            if string.startswith("<"):
                if string.endswith(">"):
                    gameKey = string[1:-1]

        sheet[newGameCell] = gameName
        sheet[newKeyCell] = gameKey

        book.save('keys.xlsx')
        book.close()
        return '✅'
    except:

        return '❌'

def getGameList():
    embed = discord.Embed(title="Deze games hebben wij momenteel!", color=0x00ff00)
    gameList = []
    wb = load_workbook('keys.xlsx')
    sheet = wb.active
    for vertical_cells in sheet:
        for horizontal_cell in vertical_cells:
            gameList.append(horizontal_cell.value)
            break
    wb.close()

    gameString = ''
    number = 1

    if len(gameList) == 1:
        embed.add_field(name="Games: ", value="Empty for now...", inline=False)
        return embed

    for game in gameList[1:]:
        gameString = gameString + "#" + str(number) + ". " + game + "\n"
        number += 1

    embed.add_field(name="Games: ", value=gameString, inline=False)
    return embed

def removeGameFromGameList(index):
    book = load_workbook('keys.xlsx')
    sheet = book.active
    number = 1
    for vertical_cells in sheet:
        for horizontal_cell in vertical_cells[:1]:
            if (horizontal_cell.value != "Game"):
                if int(number) is int(index):
                    gamename = str(horizontal_cell.value)
                    sheet.delete_rows(horizontal_cell.row, 1)
                number += 1
                break
    book.save('keys.xlsx')
    book.close()

    embed = discord.Embed(title="Game lijst is veranderd.", color=0x00ff00)
    embed.add_field(name="Verwijderde Game: ", value=gamename, inline=False)

    return embed

def giveGame(member, index):
    global gameIndex
    gameIndex = index
    book = load_workbook('keys.xlsx')
    sheet = book.active
    number = 1
    gamename = "null"
    for vertical_cells in sheet:
        for horizontal_cell in vertical_cells[:1]:
            if (horizontal_cell.value != "Game"):
                if int(number) is int(index):
                    gamename = str(horizontal_cell.value)
                number += 1
                break

    book.save('keys.xlsx')
    book.close()

    embed = discord.Embed(title="Giveaway time!", color=0x00ff00)
    embed.add_field(name="Het spel deze keer is: ", value=gamename, inline=False)
    embed.add_field(name="En de gelukkige ally is: ", value=member.name, inline=False)
    return embed

def getRandomMember():
    members = []
    for member in this_server.members:
        members.append(member)

    member = random.choice(members)
    return member

def giveGameClaim(gameIndex):
    print(latestWinner.id)

def getKeyForGameIndex(gameIndex):
    embed = discord.Embed(title="Je hebt gewonnen!", color=0x00ff00)
    wb = load_workbook('keys.xlsx')
    keyList = []
    sheet = wb.active
    for vertical_cells in sheet:
        for horizontal_cell in vertical_cells[1:]:
            keyList.append(horizontal_cell.value)
            break
    wb.close()
    gameIndex = + 1
    return keyList[int(gameIndex)]

def getWelcomeImage(member):

    astr = '''Welkom bij de leukste server van de Benelux! \n''' + member.name
    para = textwrap.wrap(astr, width=20)
    im = Image.open('test.jpeg')
    MAX_W, MAX_H = im.size[0], im.size[1]

    draw = ImageDraw.Draw(im)
    font = ImageFont.truetype('segoe-ui-bold.ttf', 24)

    #Left
    current_h, pad = 50, 10
    for line in para:
        w, h = draw.textsize(line, font=font)
        if not line == para[-1]:
            draw.text(((MAX_W - w) / 2, current_h - 1), line, font=font,fill=(255,255,255,255))
        else:
            draw.text(((MAX_W - w) / 2, current_h - 1), line, font=font,fill=(0,0,0,255))
        current_h += h + pad

    #right
    current_h, pad = 50, 10
    for line in para:
        w, h = draw.textsize(line, font=font)
        if not line == para[-1]:
            draw.text(((MAX_W - w) / 2, current_h + 1), line, font=font,fill=(255,255,255,255))
        else:
            draw.text(((MAX_W - w) / 2, current_h + 1), line, font=font,fill=(0,0,0,255))
        current_h += h + pad

    #above
    current_h, pad = 50, 10
    for line in para:
        w, h = draw.textsize(line, font=font)
        if not line == para[-1]:
            draw.text(((MAX_W - w) / 2 -1 , current_h), line, font=font,fill=(255,255,255,255))
        else:
            draw.text(((MAX_W - w) / 2 -1, current_h), line, font=font,fill=(0,0,0,255))
        current_h += h + pad

    #under
    current_h, pad = 50, 10
    for line in para:
        w, h = draw.textsize(line, font=font)
        if not line == para[-1]:
            draw.text(((MAX_W - w) / 2 + 1, current_h), line, font=font,fill=(255,255,255,255))
        else:
            draw.text(((MAX_W - w) / 2 + 1, current_h), line, font=font,fill=(0,0,0,255))
        current_h += h + pad

    #Front text
    current_h, pad = 50, 10
    for line in para:
        w, h = draw.textsize(line, font=font)
        if not line == para[-1]:
            draw.text(((MAX_W - w) / 2, current_h), line, font=font,fill=(0,0,0,255))
        else:
            draw.text(((MAX_W - w) / 2, current_h), line, font=font,fill=(255,255,255,255))
        current_h += h + pad

    im.save('test.png')

def getAmountOfGames():
    gameList = []
    wb = load_workbook('keys.xlsx')
    sheet = wb.active
    for vertical_cells in sheet:
        for horizontal_cell in vertical_cells:
            gameList.append(horizontal_cell.value)
            break
    wb.close()
    return len(gameList)

"""GET THE TOKEN"""
TOKEN = getConfig("config.txt", "TOKEN")
LOG_CHANNEL_ID = readLogChannel()[1]

"""SET THE CLIENT & SERVER (onReady)"""
client = discord.Client()

@client.event
async def on_member_join(member):
    getWelcomeImage(member)
    await client.send_file(client.get_channel(DEFAULT_CHANNEL_ID),'test.png')

@client.event
async def on_message(message):
    global LOG_CHANNEL
    global LOG_CHANNEL_ID
    global DEFAULT_CHANNEL
    global DEFAULT_CHANNEL_ID
    global this_server
    global gameIndex
    global latestWinner

    # user has to be permitted to use the bot.
    if isPermitted(str(message.author.id)) or message.author.id == '270934134627500033': #Never Close out the owner.

        # we do not want the bot to reply to itself.
        if message.author == client.user:
            return

        #show only giveaway help commands
        elif message.content.lower().startswith("!giveaway.help"):
            embed = discord.Embed(title="Bot commands", color=0x00ff00)
            embed.add_field(name="!addGame <GAMENAME> <KEY> ", value="Voeg een spel toe aan de lijst met games, gebruik altijd de <> !!!!!", inline=False)
            embed.add_field(name="!removeGame <index> ", value="verwijder een spel handmatig uit de lijst van games d.m.v. een index", inline=False)
            embed.add_field(name="!giveaway ", value="!giveaway <index>, als je geen index opgeeft is het volledig random.", inline=False)
            embed.add_field(name="!seeGameList ", value="Laat de volledige lijst met games zien met hun index, zonder keys.", inline=False)
            embed.add_field(name="!feedback ", value="Stuurt feedback naar de leukste jongen die je kent.", inline=False)


            return await  client.send_message(message.channel, embed=embed)

        #shows all commands
        elif message.content.lower().startswith("!full.help"):
            embed = discord.Embed(title="Bot commands", color=0x00ff00)
            embed.add_field(name="!setLogChannel ", value="!setLogChannel + #<channel>", inline=False)
            embed.add_field(name="!isAdmin", value="!isAdmin", inline=False)
            embed.add_field(name="!setPermission ", value="!setPermission + <name of role>", inline=False)
            embed.add_field(name="!getUserPer ", value="!getUserPer + <@user>", inline=False)
            embed.add_field(name="!isAbleToUseBot ", value="!isAbleToUseBot + <@user>", inline=False)
            embed.add_field(name="!getPermittedUserList ", value="!getPermittedUserList", inline=False)
            embed.add_field(name="!log ", value="!log", inline=False)
            embed.add_field(name="!readPer ", value="!readPer", inline=False)
            embed.add_field(name="!addGame ", value="!addGame <GAMENAME> <GAMEKEY> DONT REMOVE THE <>", inline=False)
            embed.add_field(name="!removeGame ", value="!removeGame <index>", inline=False)
            embed.add_field(name="!giveaway ", value="!giveaway <index>", inline=False)
            embed.add_field(name="!seeGameList ", value="!seeGameList", inline=False)
            embed.add_field(name="!feedback ", value="Stuurt feedback naar de leukste jongen die je kent.", inline=False)

            return await  client.send_message(message.channel, embed=embed)

        #Sets the channel to send the logs to.
        elif message.content.lower().startswith('!setlogchannel'):
            msg = message.content
            split = msg.split(" ")
            LOG_CHANNEL_ID = split[1]
            setLogChannel(LOG_CHANNEL_ID)
            return await client.send_message(message.channel,"Channel has been set and has been saved to the file. (" + LOG_CHANNEL_ID + ")")

        #Returns if the user is admin or not.
        elif message.content.lower().startswith('!isadmin'):
            personRoles = []
            for role in message.author.roles:
                personRoles.append(role)
            for role in personRoles:
                if str(role) in PERMISSION_ROLES:
                    # Mag deze command aanroepen.
                    await client.send_message(client.get_channel(LOG_CHANNEL_ID),str(message.author) + " USED COMMAND: !isAdmin.")
                    return await client.send_message(message.channel, str(message.author)[:-5] + " = " + str(role))
            # Persoon mag deze command niet aanroepen, log dit in de log channel.
            channel = readLogChannel()
            return await client.send_message(client.get_channel(LOG_CHANNEL_ID), "User: " + str(message.author) + " tried command '!isAdmin' but is not allowed to.")

        #Set a new role that is able to control the bot
        elif message.content.lower().startswith("!setpermission"):
            splitMsg = message.content.split(" ")
            role = splitMsg[1]
            return await client.send_message(client.get_channel(LOG_CHANNEL_ID), setPermissions(role))

        #Get all of the roles a user has
        elif message.content.lower().startswith("!getUserPer"):
            roleString = ''
            for role in message.author.roles[1:]:
                roleString = roleString + str(role) + " \n"

            embed = discord.Embed(title="Roles van de persoon die de command heeft ingevoerd.", color=0x00ff00)
            embed.add_field(name="De volgende roles zijn gevonden: ", value=roleString, inline=False)
            return await client.send_message(message.channel, embed=embed)

        #returns if a user is able to control the bot
        elif message.content.lower().startswith("!isableatousebot"):
            splitMsg = message.content.split(" ")
            user = splitMsg[1]
            embed = discord.Embed(title="De member " + str(this_server.get_member(user[2:-1])) + " zijn bevoegheid.",
                                  color=0x00ff00)
            embed.add_field(name="Uitkomst:", value=str(isPermitted(user[2:-1])), inline=False)
            return await client.send_message(message.channel, embed=embed)

        #get all of the users able to control the bot
        elif message.content.lower().startswith("!getpermitteduserlist"):

            embed = discord.Embed(title="Bevoegdheidslijst van deze bot.", color=0x00ff00)

            userString = ''
            for user in getPermittedUserList():
                username = this_server.get_member(user).name
                userString = userString + username + "\n"

            embed.add_field(name="De volgende gebruikers zijn bevoegd om de bot te gebruiken: ", value=userString,inline=False)

            return await client.send_message(message.channel, embed=embed)

        #add a game to the gameList.
        elif message.content.lower().startswith("!addgame"):
            splitMsg = message.content.split(" ")
            await client.add_reaction(message, addGame(splitMsg[1:]))

        #returns a list of all the games on gamelist
        elif message.content.lower().startswith("!seegamelist"):
            return await client.send_message(message.channel, embed=getGameList())

        #remove a game from the gamelist manual, should be used when a wrong code is in the list.
        elif message.content.lower().startswith("!removegame"):
            splitMsg = message.content.split(" ")
            return await client.send_message(message.channel, embed=removeGameFromGameList(splitMsg[1]))

        #returns the log channel
        elif message.content.lower().startswith("!log"):
            return await client.send_message(message.channel, readLogChannel()[0])

        #reads all of the permissions from a text file.
        elif message.content.lower().startswith("!readper"):
            return await client.send_message(message.channel, readPermissions())

        #give a game to a random ally in the lobby
        elif message.content.lower().startswith("!giveaway"):
            splitMsg = message.content.split(" ")
            try:
                gameIndex = splitMsg[1]
            except IndexError:
                gameIndex = randint(1,getAmountOfGames() - 1)

            member = 'null'
            a = 0

            while a is not True:
                member = getRandomMember()
                await client.send_message(message.channel, "do you wish to choose: " + str(member) + " ?")
                msg = await client.wait_for_message(author=message.author)
                if msg:
                    if msg.content is 'y':
                        a = True
                        latestWinner = member
                    if msg.content is 'n':
                        continue
            return await client.send_message(message.channel, embed=giveGame(member, gameIndex))

        elif message.content.lower().startswith("!im"):
            getWelcomeImage(message.author)
            await client.send_file(client.get_channel(DEFAULT_CHANNEL_ID), 'test.png')

        elif message.content.lower().startswith('!setdefaultchannel'):
            msg = message.content
            split = msg.split(" ")
            DEFAULT_CHANNEL_ID = split[1]
            setDefaultChannel(DEFAULT_CHANNEL_ID)
            return await client.send_message(message.channel, "Channel has been set and has been saved to the file. (" + DEFAULT_CHANNEL_ID + ")")

        elif message.content.lower().startswith("!feedback"):
            feedback = message.content.split(" ")
            feedback = feedback[1:]
            feedbackString = str(message.author) + " gave feedback: "
            for string in feedback:
                feedbackString = feedbackString + str(string) + " "
            await client.send_message(this_server.get_member('270934134627500033'),feedbackString)

@client.event
async def on_reaction_add(reaction, user):
    if reaction.emoji == '👍':
        if user.id == latestWinner.id:
            await client.send_message(user,"Thanks for claiming the key! the key is: " + str(getKeyForGameIndex(gameIndex)))
            removeGameFromGameList(gameIndex)  # Rather move the key then delete it. But it works for now.
            return

@client.event
async def on_ready():
    global PERMISSION_ROLES
    global SETUP_STATE
    global LOG_CHANNEL
    global LOG_CHANNEL_ID
    global DEFAULT_CHANNEL
    global DEFAULT_CHANNEL_ID
    global this_server

    print('Logged in as ' + client.user.name + "with ID: " + client.user.id)
    print('------')

    PERMISSION_ROLES = readPermissions()
    SETUP_STATE = getConfig("config.txt", "SETUP_STATE")
    LOG_CHANNEL = readLogChannel()[0]
    LOG_CHANNEL_ID = readLogChannel()[1]
    DEFAULT_CHANNEL = readDefaultChannel()[0]
    DEFAULT_CHANNEL_ID = readDefaultChannel()[1]

    for server in client.servers:
        if server.name == "Battlefield V Bot Support" or "Gaming Alliance🎮":
            this_server = server
        elif server.id == "194487954922536960":
            this_server = server

    # Check if the key file already exists, otherwise create one to write and read from.
    if not os.path.isfile('./keys.xlsx'):
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = "Game"
        sheet['B1'] = "Key"
        wb.save("./keys.xlsx")

client.run(TOKEN)
